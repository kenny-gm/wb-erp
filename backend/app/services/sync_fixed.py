"""
数据同步服务
支持：
1. 新店铺：自动同步90天历史数据（按天分批）
2. 已存在店铺：增量同步（只拉取新数据）
3. 数据去重：存在则更新，不存在则插入
"""
import hashlib
import time
import json
import logging
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from typing import Optional, Dict, Any, List
from sqlalchemy.orm import Session
from sqlalchemy import and_
from app.models.models import (
    Shop, Product, Order, OrderItem, InventorySnapshot,
    AdRecord, AdKeywordStat, SyncLog, InventoryRecord
)
from app.services.wb_api import WBAPIClient
from app.services.yandex_client import YandexClient

logger = logging.getLogger("sync")


class SyncService:
    """数据同步服务"""

    # 新店铺同步历史天数
    HISTORY_DAYS = 30

    # 每批请求间隔（秒），避免限流
    BATCH_INTERVAL = 1.0

    def __init__(self, db: Session, shop: Shop):
        self.db = db
        self.shop = shop
        self.shop_id = shop.id
        if shop.platform == "yandex":
            self.client = YandexClient(shop.api_token)
        else:
            self.client = WBAPIClient(shop.api_token)

    def is_new_shop(self) -> bool:
        """判断是否为新店铺（从未同步过）"""
        return self.shop.last_sync_at is None

    def _create_sync_log(self, sync_type: str) -> SyncLog:
        """创建同步日志"""
        sync_log = SyncLog(
            shop_id=self.shop_id,
            sync_type=sync_type,
            status="running"
        )
        self.db.add(sync_log)
        self.db.commit()
        return sync_log

    def _finish_sync_log(self, sync_log: SyncLog, success: bool, count: int = 0, message: str = ""):
        """完成同步日志"""
        sync_log.status = "success" if success else "failed"
        sync_log.records_count = count
        sync_log.message = message
        sync_log.finished_at = datetime.now(ZoneInfo("Asia/Shanghai"))
        self.db.commit()

    # ==================== 商品同步 ====================

    def sync_products(self, limit: int = 100, overwrite: bool = False) -> dict:
        """
        同步商品（WB使用Content API，Yandex使用offers接口）
        """
        if self.shop.platform == "yandex":
            return self.sync_yandex_products(overwrite=overwrite)

        sync_log = self._create_sync_log("products")

        try:
            count = 0
            updated = 0

            logger.info(f"开始同步产品，shop_id={self.shop_id}")

            # 先尝试Content API
            cards = self.client.get_products(limit=limit, offset=0, locale="ru")
            logger.info(f"Content API返回: {len(cards)}个产品")

            # 如果Content API返回空，使用Statistics API
            if not cards:
                logger.info("Content API返回空，尝试Statistics API")
                cards = self.client.get_products_from_statistics()
                logger.info(f"Statistics API返回: {len(cards)}个产品")

            if not cards:
                logger.info("没有产品数据")
                self._finish_sync_log(sync_log, True, 0, "没有产品数据")
                return {"success": True, "count": 0, "updated": 0}

            # 打印第一个产品的结构用于调试
            logger.info(f"第一个产品数据示例: {json.dumps(cards[0], ensure_ascii=False)[:300]}")

            for card in cards:
                # 根据API来源解析字段
                nm_id = str(card.get("nmID", card.get("nmId", "")))
                if not nm_id:
                    logger.warning(f"产品缺少nmID: {card}")
                    continue

                # vendorCode 或 supplierArticle 是 SKU
                vendor_code = card.get("vendorCode", card.get("supplierArticle", ""))

                # 获取产品名称
                name = f"产品 {nm_id}"
                characteristics = card.get("characterstics", [])
                for char in characteristics:
                    if char.get("name") == "Наименование" or char.get("id") == 0:
                        name = char.get("value", "")
                        break

                # 如果没有从characteristics获取到名称，使用默认名称
                if name == f"产品 {nm_id}":
                    # Statistics API返回的数据没有characteristics，使用SKU作为名称
                    if vendor_code:
                        name = vendor_code

                logger.info(f"处理产品: nmID={nm_id}, vendorCode={vendor_code}, name={name[:30]}")

                # 查找已存在商品
                existing = self.db.query(Product).filter(
                    Product.nm_id == nm_id,
                    Product.shop_id == self.shop_id
                ).first()

                # 提取尺寸信息
                dimensions = card.get("dimensions", {})

                if existing:
                    if overwrite:
                        # 更新已存在商品
                        existing.name = name
                        existing.sku = vendor_code
                        existing.weight = dimensions.get("weightBrutto", 0)
                        existing.length = dimensions.get("length", 0)
                        existing.width = dimensions.get("width", 0)
                        existing.height = dimensions.get("height", 0)
                        updated += 1
                        logger.info(f"更新产品: {nm_id}")
                    continue
                else:
                    # 检查nm_id是否已在其他店铺存在（nm_id全局唯一）
                    existing_global = self.db.query(Product).filter(
                        Product.nm_id == nm_id
                    ).first()
                    if existing_global:
                        # nm_id已存在，直接跳过（不重复创建）
                        logger.info(f"nmID={nm_id}已存在于店铺{existing_global.shop_id}，跳过")
                        continue
                    # 提取尺寸信息
                    dimensions = card.get("dimensions", {})

                    # 创建新商品
                    product = Product(
                        nm_id=nm_id,
                        sku=vendor_code,
                        shop_id=self.shop_id,
                        name=name,
                        custom_name=name,
                        weight=dimensions.get("weightBrutto", 0),
                        length=dimensions.get("length", 0),
                        width=dimensions.get("width", 0),
                        height=dimensions.get("height", 0),
                    )
                    self.db.add(product)
                    count += 1
                    logger.info(f"新增产品: nmID={nm_id}, name={name[:30]}")

            # 更新同步时间
            self.shop.last_sync_at = datetime.now(ZoneInfo("Asia/Shanghai"))
            self.db.commit()

            self._finish_sync_log(sync_log, True, count, f"新增 {count} 个，更新 {updated} 个")

            return {"success": True, "count": count, "updated": updated}

        except Exception as e:
            logger.error(f"同步商品失败: {e}")
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}

    # ==================== 订单同步 ====================

    def sync_orders(self, days: Optional[int] = 30, incremental: bool = True) -> dict:
        """
        同步订单（WB使用analytics，Yandex使用stats/orders）
        """
        if self.shop.platform == "yandex":
            return self.sync_yandex_orders(days=days)

        sync_log = self._create_sync_log("orders")

        try:
            products = self.db.query(Product).filter(
                Product.shop_id == self.shop_id,
                Product.nm_id.isnot(None)
            ).all()

            nm_ids = [int(p.nm_id) for p in products if p.nm_id and p.nm_id.isdigit()]

            if not nm_ids:
                self._finish_sync_log(sync_log, True, 0, "没有产品")
                return {"success": True, "count": 0, "updated": 0}

            # 限制为7天
            if days is None or days > 7:
                days = 7
            date_to = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d")
            date_from = (datetime.now(ZoneInfo("Asia/Shanghai")) - timedelta(days=days)).strftime("%Y-%m-%d")

            # 获取产品销售漏斗数据
            analytics_result = self.client.get_product_sales_funnel(nm_ids, date_from, date_to)

            if not analytics_result:
                self._finish_sync_log(sync_log, True, 0, "无数据")
                return {"success": True, "count": 0, "updated": 0}

            # 聚合每日销售数据 (analytics_result格式: {nm_id: {date: data}})
            daily_sales = {}
            for nm_id, dates_data in analytics_result.items():
                for date_str, day_data in dates_data.items():
                    if date_str not in daily_sales:
                        daily_sales[date_str] = {"order_count": 0, "order_sum": 0}
                    daily_sales[date_str]["order_count"] += day_data.get("order_count", 0)
                    daily_sales[date_str]["order_sum"] += day_data.get("order_sum", 0)

            count = 0
            updated = 0

            for date_str, day_data in daily_sales.items():
                try:
                    order_date = datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    continue

                total_amount = day_data.get("order_sum", 0)
                order_count = day_data.get("order_count", 0)

                if order_count == 0:
                    continue

                order_id = f"analytics_{self.shop_id}_{date_str}"

                existing = self.db.query(Order).filter(
                    Order.order_id == order_id,
                    Order.shop_id == self.shop_id
                ).first()

                if existing:
                    existing.total_amount = total_amount
                    existing.updated_at = datetime.now(ZoneInfo("Asia/Shanghai"))
                    updated += 1
                else:
                    order = Order(
                        order_id=order_id,
                        shop_id=self.shop_id,
                        status="new",
                        total_amount=total_amount,
                        commission=0,
                        logistics_fee=0,
                        order_date=order_date
                    )
                    self.db.add(order)
                    count += 1

            self.db.commit()
            self._finish_sync_log(sync_log, True, count, f"新增 {count} 条，更新 {updated} 条")
            return {"success": True, "count": count, "updated": updated}

        except Exception as e:
            self.db.rollback()
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}


    def sync_inventory(self) -> dict:
        if self.shop.platform == "yandex":
            return {"success": True, "message": "Yandex MVP 暂不支持库存"}
        sync_log = self._create_sync_log("inventory")
        try:
            self._finish_sync_log(sync_log, True, 0, "跳过库存同步")
            return {"success": True, "count": 0}
        except Exception as e:
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}


    def sync_ads(self, days: Optional[int] = 30) -> dict:
        """
        同步广告数据 - 使用 fullstats API
        """
        sync_log = self._create_sync_log("ads")

        try:
            date_to = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d")

            if days:
                date_from = (datetime.now(ZoneInfo("Asia/Shanghai")) - timedelta(days=days)).strftime("%Y-%m-%d")
            elif self.shop.last_sync_at:
                date_from = (self.shop.last_sync_at - timedelta(days=1)).strftime("%Y-%m-%d")
            else:
                date_from = (datetime.now(ZoneInfo("Asia/Shanghai")) - timedelta(days=days)).strftime("%Y-%m-%d")

            stats = self.client.get_ad_stats(date_from=date_from, date_to=date_to)

            # Step 0: 获取广告配置信息
            adverts = self.client.get_adverts()
            advert_info_map = {}
            for ad in adverts:
                ad_id = ad.get("id")
                if not ad_id:
                    continue
                settings = ad.get("settings", {})
                payment_type = settings.get("payment_type", "") or ""
                placements_dict = settings.get("placements", {})
                is_search = placements_dict.get("search", False) if isinstance(placements_dict, dict) else False
                is_rec = placements_dict.get("recommendations", False) if isinstance(placements_dict, dict) else False
                if is_search and is_rec:
                    placements = "search+rec"
                elif is_search:
                    placements = "search"
                elif is_rec:
                    placements = "recommendations"
                else:
                    placements = ""
                advert_info_map[ad_id] = {
                    "payment_type": payment_type,
                    "placements": placements,
                }
            logger.info(f"获取到 {len(advert_info_map)} 个广告配置")

            # Step 1: aggregate by (advert_id, nm_id, record_date)
            aggregated = {}

            for stat in stats:
                advert_id = stat.get("advertId")
                days_data = stat.get("days", [])

                for day_data in days_data:
                    date_str = day_data.get("date", "")
                    if date_str:
                        try:
                            record_date = datetime.fromisoformat(date_str.replace("Z", "+00:00")).replace(tzinfo=None)
                        except:
                            record_date = datetime.now(ZoneInfo("Asia/Shanghai"))
                    else:
                        record_date = datetime.now(ZoneInfo("Asia/Shanghai"))

                    apps = day_data.get("apps", [])
                    for app in apps:
                        nms = app.get("nms", [])
                        for nm_data in nms:
                            nm_id = str(nm_data.get("nmId", ""))
                            if not nm_id:
                                continue

                            nm_sum = nm_data.get("sum", 0) or 0
                            nm_views = nm_data.get("views", 0) or 0
                            nm_clicks = nm_data.get("clicks", 0) or 0
                            nm_orders = nm_data.get("orders", 0) or 0
                            nm_atbs = nm_data.get("atbs", 0) or 0
                            nm_shks = nm_data.get("shks", 0) or 0

                            key = (advert_id, nm_id, record_date)
                            if key not in aggregated:
                                aggregated[key] = {"sum": 0, "views": 0, "clicks": 0, "orders": 0, "atbs": 0, "shks": 0}
                            aggregated[key]["sum"] += nm_sum
                            aggregated[key]["views"] += nm_views
                            aggregated[key]["clicks"] += nm_clicks
                            aggregated[key]["orders"] += nm_orders
                            aggregated[key]["atbs"] += nm_atbs
                            aggregated[key]["shks"] += nm_shks

            # Step 2: write to DB
            count = 0
            updated = 0

            for (advert_id, nm_id, record_date), data in aggregated.items():
                product = self.db.query(Product).filter(
                    Product.nm_id == nm_id,
                    Product.shop_id == self.shop_id
                ).first()

                if not product:
                    continue

                ad_info = advert_info_map.get(advert_id, {})
                ad_payment_type = ad_info.get("payment_type", "") or ""
                ad_placements = ad_info.get("placements", "") or ""
                visitors = data["clicks"]

                existing = self.db.query(AdRecord).filter(
                    AdRecord.shop_id == self.shop_id,
                    AdRecord.product_id == product.id,
                    AdRecord.record_date == record_date,
                    AdRecord.ad_type == "advertising",
                    AdRecord.advert_id == advert_id
                ).first()

                if existing:
                    existing.impressions = data["views"]
                    existing.clicks = data["clicks"]
                    existing.visitors = visitors
                    existing.cost = data["sum"]
                    existing.order_count = data["orders"]
                    existing.payment_type = ad_payment_type
                    existing.placements = ad_placements
                    existing.atbs = data.get("atbs", 0)
                    existing.shks = data.get("shks", 0)
                    existing.platform = ""
                    existing.cart_count = data.get("atbs", 0)
                    updated += 1
                else:
                    ad_record = AdRecord(
                        product_id=product.id,
                        shop_id=self.shop_id,
                        ad_type="advertising",
                        advert_id=advert_id,
                        platform="",
                        impressions=data["views"],
                        clicks=data["clicks"],
                        visitors=visitors,
                        cost=data["sum"],
                        order_count=data["orders"],
                        record_date=record_date,
                        payment_type=ad_payment_type,
                        placements=ad_placements,
                        atbs=data.get("atbs", 0),
                        shks=data.get("shks", 0),
                        cart_count=data.get("atbs", 0)
                    )
                    self.db.add(ad_record)
                    count += 1

            self.shop.last_sync_at = datetime.now(ZoneInfo("Asia/Shanghai"))
            self.db.commit()

            self._finish_sync_log(sync_log, True, count, f"新增 {count} 条广告记录")

            return {"success": True, "count": count, "updated": updated}

        except Exception as e:
            self.db.rollback()
            logger.error(f"同步广告失败: {e}")
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}

    # ==================== 产品销售漏斗 ====================

    def sync_product_sales(self, days: Optional[int] = 30) -> dict:
        """
        同步产品销售漏斗数据（WB用analytics，Yandex已整合到sync_yandex_orders）
        """
        if self.shop.platform == "yandex":
            return self.sync_yandex_product_sales(days)

        sync_log = self._create_sync_log("product_sales")

        try:
            products = self.db.query(Product).filter(
                Product.shop_id == self.shop_id,
                Product.nm_id != "0"
            ).all()

            if not products:
                self._finish_sync_log(sync_log, True, 0, "无产品需要同步")
                return {"success": True, "count": 0, "updated": 0}

            if days is None or days > 7:
                days = 7

            date_to = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d")
            date_from = (datetime.now(ZoneInfo("Asia/Shanghai")) - timedelta(days=days)).strftime("%Y-%m-%d")

            nm_ids = [int(p.nm_id) for p in products if p.nm_id and p.nm_id.isdigit()]

            if not nm_ids:
                self._finish_sync_log(sync_log, True, 0, "无有效产品ID")
                return {"success": True, "count": 0, "updated": 0}

            logger.info(f"开始同步产品销售数据: {len(nm_ids)}个产品, {date_from} - {date_to}")

            product_sales_data = {}
            batch_size = 20
            for i in range(0, len(nm_ids), batch_size):
                batch_nm_ids = nm_ids[i:i+batch_size]
                batch_result = self.client.get_product_sales_funnel(batch_nm_ids, date_from, date_to)
                if batch_result:
                    product_sales_data.update(batch_result)

            if not product_sales_data:
                self._finish_sync_log(sync_log, True, 0, "API返回空数据")
                return {"success": True, "count": 0, "updated": 0}

            nm_to_product = {p.nm_id: p.id for p in products}

            count = 0
            updated = 0

            for nm_id_str, daily_data in product_sales_data.items():
                product_id = nm_to_product.get(nm_id_str)
                if not product_id:
                    continue

                for date_str, stats in daily_data.items():
                    try:
                        record_date = datetime.strptime(date_str, "%Y-%m-%d")
                    except:
                        continue

                    existing = self.db.query(AdRecord).filter(
                        AdRecord.record_date == record_date,
                        AdRecord.ad_type == "product_analytics",
                        AdRecord.product_id == product_id
                    ).first()

                    order_sum = stats.get("order_sum", 0)
                    if existing:
                        existing.impressions = 0
                        existing.visitors = stats.get("visitors", 0)
                        existing.cart_count = stats.get("cart_count", 0)
                        existing.order_count = stats.get("order_count", 0)
                        existing.sales = order_sum
                        updated += 1
                    else:
                        ad_record = AdRecord(
                            shop_id=self.shop_id,
                            product_id=product_id,
                            ad_type="product_analytics",
                            impressions=0,
                            visitors=stats.get("visitors", 0),
                            clicks=0,
                            cart_count=stats.get("cart_count", 0),
                            order_count=stats.get("order_count", 0),
                            sales=order_sum,
                            record_date=record_date
                        )
                        self.db.add(ad_record)
                        count += 1

            self.db.commit()

            total = count + updated
            logger.info(f"产品销售数据同步完成: 新增{count}条, 更新{updated}条, 总计{total}条")

            self._finish_sync_log(sync_log, True, total, f"同步{total}条产品销售数据")
            return {"success": True, "count": count, "updated": updated, "total": total}

        except Exception as e:
            logger.error(f"同步产品销售数据失败: {e}")
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}

    # ============================================================
    # Yandex 同步方法
    # ============================================================

    def sync_yandex_products(self, overwrite: bool = False) -> dict:
        """
        Yandex 商品同步：遍历 business 下所有 campaign 的 offers，按 offer_id 去重写入 Product。
        成功写入后更新 last_sync_at，失败时不更新。
        """
        sync_log = self._create_sync_log("products")

        try:
            config = self.shop.platform_config or {}
            business_id = config.get("business_id")
            campaign_ids = config.get("campaign_ids", [])

            if not business_id:
                msg = "Yandex products: platform_config 缺少 business_id"
                logger.error(msg)
                self._finish_sync_log(sync_log, False, 0, msg)
                return {"success": False, "error": msg}
            if not campaign_ids:
                msg = "Yandex products: platform_config campaign_ids 为空"
                logger.error(msg)
                self._finish_sync_log(sync_log, False, 0, msg)
                return {"success": False, "error": msg}
            logger.info(f"Yandex products: business={business_id}, campaigns={campaign_ids}")

            offers = self.client.get_all_offers(int(business_id), campaign_ids)

            if not offers:
                # API 成功但无商品数据，也更新 last_sync_at
                self.shop.last_sync_at = datetime.now(ZoneInfo("Asia/Shanghai"))
                self.db.commit()
                self._finish_sync_log(sync_log, True, 0, "无商品数据")
                return {"success": True, "count": 0, "updated": 0}

            count = 0
            updated = 0

            for offer in offers:
                oid = offer["offer_id"]
                offer_name = offer.get("offer_name") or oid

                product = self.db.query(Product).filter(
                    Product.shop_id == self.shop_id,
                    Product.sku == oid
                ).first()

                if product:
                    if overwrite and offer_name and offer_name != oid:
                        product.name = offer_name
                    updated += 1
                else:
                    nm_id = "ym_" + hashlib.sha1(
                        f"{int(business_id)}:{oid}".encode()
                    ).hexdigest()[:32]
                    product = Product(
                        nm_id=nm_id,
                        sku=oid,
                        shop_id=self.shop_id,
                        name=offer_name,
                        custom_name=offer_name,
                    )
                    self.db.add(product)
                    count += 1

            self.db.commit()

            # 成功后才更新 last_sync_at
            self.shop.last_sync_at = datetime.now(ZoneInfo("Asia/Shanghai"))
            self.db.commit()

            total = count + updated
            logger.info(f"Yandex products 完成: 新增 {count}, 更新 {updated}")
            self._finish_sync_log(sync_log, True, total, f"新增 {count} 个，更新 {updated} 个")
            return {"success": True, "count": count, "updated": updated}

        except Exception as e:
            logger.error(f"Yandex products 同步异常: {e}", exc_info=True)
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}

    def sync_yandex_orders(self, days: Optional[int] = 7) -> dict:
        """
        Yandex 订单同步：
        1. 遍历 business 下所有 campaign 的 stats/orders
        2. 写入 Order / OrderItem（保留订单明细）
        3. 同时按 product_id + day 聚合写入 AdRecord(product_analytics)

        MVP 阶段：访客/点击/加购/广告费均写 0
        """
        sync_log = self._create_sync_log("orders")

        try:
            config = self.shop.platform_config or {}
            business_id = config.get("business_id")
            campaign_ids = config.get("campaign_ids", [])

            if not business_id:
                msg = "Yandex orders: platform_config 缺少 business_id"
                logger.error(msg)
                self._finish_sync_log(sync_log, False, 0, msg)
                return {"success": False, "error": msg}
            if not campaign_ids:
                msg = "Yandex orders: platform_config campaign_ids 为空"
                logger.error(msg)
                self._finish_sync_log(sync_log, False, 0, msg)
                return {"success": False, "error": msg}

            # Yandex stats/orders 最大 7 天
            if not days or days > 7:
                days = 7

            date_to = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d")
            date_from = (
                datetime.now(ZoneInfo("Asia/Shanghai")) - timedelta(days=days)
            ).strftime("%Y-%m-%d")

            logger.info(
                f"Yandex orders: business={business_id}, campaigns={campaign_ids}, "
                f"{date_from} - {date_to}"
            )

            # -----------------------------------------------------------
            # Step 1: 写入 Order / OrderItem
            # -----------------------------------------------------------
            order_rows = self.client.get_orders_for_db(
                int(business_id), campaign_ids, date_from, date_to
            )

            order_count = 0
            order_updated = 0
            item_count = 0

            # 按 order_id 聚合，同时记录 campaign_id（用于 order_id 命名空间）
            # items 用 Dict[offer_id, row] 去重：同一订单内同 offer_id 只保留一条（累加数量）
            order_map: Dict[str, dict] = {}
            for row in order_rows:
                oid = row["order_id"]
                if oid not in order_map:
                    order_map[oid] = {
                        "order_id": oid,
                        "order_date": row["order_date"],
                        "campaign_id": row["campaign_id"],
                        "total_amount": 0.0,
                        "items": {},  # Dict[offer_id, row] 用于同订单内同 sku 去重
                    }
                order_map[oid]["total_amount"] += row["total_price"]
                # 同订单内同 offer_id 累加数量（不去重则双计）
                existing_row = order_map[oid]["items"].get(row["offer_id"])
                if existing_row:
                    existing_row["quantity"] += row["quantity"]
                    existing_row["total_price"] += row["total_price"]
                else:
                    order_map[oid]["items"][row["offer_id"]] = row

            for oid, odata in order_map.items():
                try:
                    order_date = datetime.strptime(odata["order_date"], "%Y-%m-%d")
                except ValueError:
                    order_date = datetime.now(ZoneInfo("Asia/Shanghai"))

                # Yandex order_id 命名空间：yandex:{business_id}:{campaign_id}:{order_id}
                namespaced_order_id = f"yandex:{business_id}:{odata['campaign_id']}:{oid}"

                existing_order = self.db.query(Order).filter(
                    Order.order_id == namespaced_order_id,
                    Order.shop_id == self.shop_id
                ).first()

                if existing_order:
                    existing_order.total_amount = odata["total_amount"]
                    existing_order.updated_at = datetime.now(ZoneInfo("Asia/Shanghai"))
                    order_updated += 1
                else:
                    order = Order(
                        order_id=namespaced_order_id,
                        shop_id=self.shop_id,
                        status="new",
                        total_amount=odata["total_amount"],
                        order_date=order_date,
                    )
                    self.db.add(order)
                    self.db.flush()
                    order_count += 1

                order_obj = existing_order or order

                for item in odata["items"].values():
                    product = self.db.query(Product).filter(
                        Product.shop_id == self.shop_id,
                        Product.sku == item["offer_id"]
                    ).first()

                    existing_item = self.db.query(OrderItem).filter(
                        OrderItem.order_id == order_obj.id,
                        OrderItem.sku == item["offer_id"]
                    ).first()

                    if not existing_item:
                        order_item = OrderItem(
                            order_id=order_obj.id,
                            product_id=product.id if product else None,
                            nm_id=product.nm_id if product else "",
                            sku=item["offer_id"],
                            quantity=item["quantity"],
                            price=item["price"],
                            total_price=item["total_price"],
                        )
                        self.db.add(order_item)
                        item_count += 1
                    else:
                        # 幂等更新：重复同步同一订单时更新数量和价格
                        existing_item.quantity = item["quantity"]
                        existing_item.price = item["price"]
                        existing_item.total_price = item["total_price"]

            logger.info(
                f"Yandex orders: 新增 Order={order_count}, 更新 Order={order_updated}, "
                f"新增 OrderItem={item_count}"
            )

            # -----------------------------------------------------------
            # Step 2: 按 product_id + day 聚合写入 AdRecord(product_analytics)
            # -----------------------------------------------------------
            aggregated = self.client.get_orders_aggregated(
                int(business_id), campaign_ids, date_from, date_to
            )

            ad_count = 0
            ad_updated = 0

            # 按 (offer_id, day) 聚合
            agg_map: Dict[tuple, dict] = {}
            for row in aggregated:
                key = (row["offer_id"], row["day"])
                if key not in agg_map:
                    agg_map[key] = {
                        "offer_id": row["offer_id"],
                        "offer_name": row.get("offer_name") or "",
                        "day": row["day"],
                        "order_count": 0,
                        "sales_amount": 0.0,
                    }
                agg_map[key]["order_count"] += row["order_count"]
                agg_map[key]["sales_amount"] += row["sales_amount"]

            for (offer_id, day_str), data in agg_map.items():
                product = self.db.query(Product).filter(
                    Product.shop_id == self.shop_id,
                    Product.sku == offer_id
                ).first()

                if not product:
                    # 自动创建 Product（从订单发现新商品）
                    nm_id = "ym_" + hashlib.sha1(
                        f"{int(business_id)}:{offer_id}".encode()
                    ).hexdigest()[:32]
                    product = Product(
                        nm_id=nm_id,
                        sku=offer_id,
                        shop_id=self.shop_id,
                        name=data["offer_name"] or offer_id,
                        custom_name=data["offer_name"] or offer_id,
                    )
                    self.db.add(product)
                    self.db.flush()

                try:
                    record_date = datetime.strptime(day_str, "%Y-%m-%d")
                except ValueError:
                    continue

                existing = self.db.query(AdRecord).filter(
                    AdRecord.shop_id == self.shop_id,
                    AdRecord.product_id == product.id,
                    AdRecord.record_date == record_date,
                    AdRecord.ad_type == "product_analytics"
                ).first()
                if existing:
                    # 只更新订单和销售额字段，不清零已有流量字段
                    # （流量由 traffic 同步写入，orders 同步不能覆盖）
                    # existing.visitors = 0
                    # existing.clicks = 0
                    # existing.cart_count = 0
                    # existing.impressions = 0
                    existing.order_count = data["order_count"]
                    existing.sales = data["sales_amount"]
                    existing.cost = 0
                    ad_updated += 1
                else:
                    ad_record = AdRecord(
                        shop_id=self.shop_id,
                        product_id=product.id,
                        ad_type="product_analytics",
                        impressions=0,
                        visitors=0,
                        clicks=0,
                        cart_count=0,
                        order_count=data["order_count"],
                        sales=data["sales_amount"],
                        cost=0,
                        record_date=record_date,
                    )
                    self.db.add(ad_record)
                    ad_count += 1

            self.db.commit()

            # 成功后才更新 last_sync_at
            self.shop.last_sync_at = datetime.now(ZoneInfo("Asia/Shanghai"))
            self.db.commit()

            total = order_count + order_updated + item_count + ad_count + ad_updated
            logger.info(
                f"Yandex orders 完成: Order 新增 {order_count} 更新 {order_updated}, "
                f"OrderItem 新增 {item_count}, "
                f"AdRecord 新增 {ad_count} 更新 {ad_updated}"
            )
            self._finish_sync_log(
                sync_log, True, total,
                f"Order {order_count}/{order_updated}, Item {item_count}, "
                f"AdRecord 新增 {ad_count} 更新 {ad_updated}"
            )
            return {
                "success": True,
                "order_count": order_count,
                "order_updated": order_updated,
                "item_count": item_count,
                "ad_count": ad_count,
                "ad_updated": ad_updated,
            }

        except Exception as e:
            logger.error(f"Yandex orders 同步异常: {e}", exc_info=True)
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}

    def sync_yandex_product_sales(self, days: Optional[int] = 7) -> dict:
        """
        Yandex product_sales 已由 sync_yandex_orders 整合写入 AdRecord。
        这里只返回提示信息，不重复同步。
        """
        return {
            "success": True,
            "message": "Yandex product_sales 已由 orders 同步生成，无需重复拉取"
        }

    def sync_yandex_traffic(self, date_from: Optional[str] = None, date_to: Optional[str] = None) -> dict:
        """
        Yandex 流量数据同步（shows-sales 报告）：
        1. 优先使用 businessId 级别请求（一次只生成1个报表，覆盖所有 campaign）
        2. 轮询直到报告生成完成（异步，约5-10分钟）
        3. 下载 XLSX，解析 "Аналитика продаж" 工作表
        4. 按 offer_id + date 聚合后写入 AdRecord(product_analytics) 流量字段

        只允许写入流量字段：impressions / visitors / cart_count / clicks
        严禁写入：order_count / sales / cost
        """
        import httpx
        import io
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        sync_log = self._create_sync_log("traffic")

        try:
            config = self.shop.platform_config or {}
            campaign_ids = config.get("campaign_ids", [])
            business_id = config.get("business_id")

            if not business_id:
                msg = "Yandex traffic: platform_config 缺少 business_id"
                logger.error(msg)
                self._finish_sync_log(sync_log, False, 0, msg)
                return {"success": False, "error": msg}

            # 默认同步昨天（北京时间）
            if not date_from or not date_to:
                tz = ZoneInfo("Asia/Shanghai")
                yesterday = (datetime.now(tz) - timedelta(days=1)).strftime("%Y-%m-%d")
                date_from = yesterday
                date_to = yesterday

            logger.info(f"Yandex traffic: businessId={business_id}, date={date_from} to {date_to}")

            # 优先使用 businessId 级别请求（单次请求，覆盖所有 campaign）
            all_records = []
            business_request_succeeded = False

            if business_id:
                try:
                    result = self._fetch_shows_sales_report_business(business_id, date_from, date_to)

                    if result.get("rate_limited"):
                        logger.warning(f"  businessId shows-sales 420 rate limit，直接返回")
                        self._finish_sync_log(sync_log, False, 0, "Yandex shows-sales rate limited, retry later")
                        return {
                            "success": False,
                            "rate_limited": True,
                            "retry_after_seconds": 600,
                            "message": "Yandex shows-sales rate limited, retry later"
                        }

                    if result.get("error") and not result.get("fallback_allowed"):
                        # 非420失败，不fallback，直接返回错误
                        logger.warning(f"  businessId 请求失败（不fallback）: {result['error']}")
                        self._finish_sync_log(sync_log, False, 0, result["error"])
                        return {"success": False, "error": result["error"]}

                    if result.get("success"):
                        business_request_succeeded = True
                        all_records = result.get("records", [])
                        logger.info(f"  businessId {business_id}: 获取 {len(all_records)} 条流量记录（请求成功，不fallback）")
                    elif result.get("fallback_allowed"):
                        # 只有 fallback_allowed=True（400参数错误）才走fallback
                        logger.warning(f"  businessId 请求失败，允许fallback: {result.get('error', 'unknown')}")
                        business_request_succeeded = False
                except Exception as e:
                    logger.warning(f"  businessId 请求异常，fallback到campaign: {e}")
                    business_request_succeeded = False

            # Fallback 只允许：400 参数错误导致 fallback_allowed=True
            # 420 / 超时 / 解析失败 / 缺sheet / 无reportId → 不fallback，直接返回错误
            if not all_records and not business_request_succeeded and campaign_ids:
                logger.info(f"  启用 campaign 串行 fallback（仅限 businessId 请求失败）")
                for cid in campaign_ids:
                    try:
                        records = self._fetch_shows_sales_report(cid, date_from, date_to)
                        all_records.extend(records)
                        logger.info(f"  campaign {cid}: 获取 {len(records)} 条流量记录")
                    except httpx.HTTPStatusError as e:
                        if e.response.status_code == 420:
                            # 420 不重试，不等待，直接返回 rate_limited
                            logger.warning(f"  campaign {cid}: 420 rate limit，不继续请求，直接返回")
                            self._finish_sync_log(sync_log, False, 0, "Yandex shows-sales rate limited, retry later")
                            return {
                                "success": False,
                                "rate_limited": True,
                                "retry_after_seconds": 600,
                                "message": "Yandex shows-sales rate limited, retry later"
                            }
                        else:
                            raise

            if not all_records:
                self._finish_sync_log(sync_log, True, 0, "无流量数据")
                return {"success": True, "count": 0}

            # 按 offer_id + date 聚合（合并多个 campaign 的同商品流量）
            aggregated: Dict[tuple, dict] = {}
            for rec in all_records:
                key = (rec["offer_id"], rec["date"][:10])
                if key not in aggregated:
                    aggregated[key] = {
                        "offer_id": rec["offer_id"],
                        "date": rec["date"][:10],
                        "shows": 0,
                        "clicks": 0,
                        "to_cart": 0,
                    }
                aggregated[key]["shows"] += rec.get("shows", 0) or 0
                aggregated[key]["clicks"] += rec.get("clicks", 0) or 0
                aggregated[key]["to_cart"] += rec.get("to_cart", 0) or 0

            aggregated_list = list(aggregated.values())
            logger.info(f"  聚合后: {len(aggregated_list)} 条唯一 (offer_id, date) 记录")

            # 写入 AdRecord（只写流量字段）
            count = 0
            for rec in aggregated_list:
                sku = rec["offer_id"]
                product = self.db.query(Product).filter(
                    Product.shop_id == self.shop_id,
                    Product.sku == sku
                ).first()
                if not product:
                    logger.warning(f"  Yandex traffic: 产品 SKU={sku} 不存在，跳过")
                    continue

                record_date = datetime.strptime(rec["date"][:10], "%Y-%m-%d").date()
                existing = self.db.query(AdRecord).filter(
                    AdRecord.shop_id == self.shop_id,
                    AdRecord.product_id == product.id,
                    AdRecord.record_date == record_date,
                    AdRecord.ad_type == "product_analytics"
                ).first()
                if existing:
                    # 只更新流量字段，不触碰 order_count/sales
                    existing.impressions = rec.get("shows", 0) or 0
                    existing.visitors = rec.get("clicks", 0) or 0
                    existing.cart_count = rec.get("to_cart", 0) or 0
                else:
                    ad_record = AdRecord(
                        shop_id=self.shop_id,
                        product_id=product.id,
                        record_date=record_date,
                        ad_type="product_analytics",
                        impressions=rec.get("shows", 0) or 0,
                        visitors=rec.get("clicks", 0) or 0,
                        cart_count=rec.get("to_cart", 0) or 0,
                        order_count=0,
                        sales=0.0,
                    )
                    self.db.add(ad_record)
                    count += 1

            self.db.commit()
            self.shop.last_sync_at = datetime.now(ZoneInfo("Asia/Shanghai"))
            self.db.commit()

            logger.info(f"Yandex traffic 完成: 新增 {count} 条")
            self._finish_sync_log(sync_log, True, count, f"新增 {count} 条")
            return {"success": True, "count": count}

        except Exception as e:
            logger.error(f"Yandex traffic 同步异常: {e}", exc_info=True)
            self._finish_sync_log(sync_log, False, 0, str(e))
            return {"success": False, "error": str(e)}

    def _fetch_shows_sales_report_business(self, business_id: int, date_from: str, date_to: str) -> dict:
        """
        返回结构化 dict:
          - 200成功有数据: {success=True, records=[...], rate_limited=False, fallback_allowed=False, error=null}
          - 200成功无数据: {success=True, records=[], rate_limited=False, fallback_allowed=False, error=null}
          - 420: {success=False, records=[], rate_limited=True, fallback_allowed=False, error="420"}
          - 400: {success=False, records=[], rate_limited=False, fallback_allowed=True, error="400"}
          - 其他失败: {success=False, records=[], rate_limited=False, fallback_allowed=False, error=具体信息}
        """
        import httpx
        import io
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        headers = {"Api-Key": self.client.api_token}

        # Step 1: 发起 business 级别报告生成
        body = {
            "businessId": business_id,
            "dateFrom": date_from,
            "dateTo": date_to,
            "grouping": "OFFERS"
        }
        logger.info(f"  请求 businessId={business_id} shows-sales 报表...")
        resp = httpx.post(
            "https://api.partner.market.yandex.ru/v2/reports/shows-sales/generate",
            headers=headers, json=body, timeout=30
        )

        # 420: rate limit，request_failed=False（由调用方处理，不走 fallback）
        if resp.status_code == 420:
            logger.warning(f"  businessId shows-sales 420 rate limit")
            return {"success": False, "records": [], "rate_limited": True, "fallback_allowed": False, "error": "420 rate limit"}

        # 400: 参数错误，request_failed=True（允许 fallback）
        if resp.status_code == 400:
            logger.warning(f"  businessId 请求返回 400（参数错误）: {resp.text}")
            return {"success": False, "records": [], "rate_limited": False, "fallback_allowed": True, "error": f"400 parameter error: {resp.text}"}

        resp.raise_for_status()
        data = resp.json()
        report_id = data.get("result", {}).get("reportId")
        if not report_id:
            logger.error(f"shows-sales 生成失败: {data}")
            return {"success": False, "records": [], "rate_limited": False, "fallback_allowed": False, "error": "shows-sales reportId missing"}

        # Step 2: 轮询等待（最多15分钟）
        for i in range(90):
            time.sleep(10)
            info_resp = httpx.get(
                f"https://api.partner.market.yandex.ru/v2/reports/info/{report_id}",
                headers=headers, timeout=15
            )
            info = info_resp.json()
            status = info.get("result", {}).get("status")
            logger.info(f"  shows-sales poll {i+1}: {status}")
            if status == "DONE":
                break
            if status == "FAILED":
                logger.error(f"shows-sales 报告生成失败: {info}")
                return {"success": False, "records": [], "rate_limited": False, "fallback_allowed": False, "error": "shows-sales report FAILED"}
        else:
            logger.error("shows-sales 报告生成超时（15分钟）")
            return {"success": False, "records": [], "rate_limited": False, "fallback_allowed": False, "error": "shows-sales report timeout (15min)"}

        # Step 3: 下载并解析 XLSX
        download_url = info["result"]["file"]
        dl_resp = httpx.get(download_url, timeout=60)
        dl_resp.raise_for_status()

        try:
            wb = load_workbook(io.BytesIO(dl_resp.content))
        except InvalidFileException:
            logger.error("报告文件解析失败（非 XLSX 格式）")
            return {"success": False, "records": [], "rate_limited": False, "fallback_allowed": False, "error": "report file parse failed"}

        sheet_name = "Аналитика продаж"
        if sheet_name not in wb.sheetnames:
            logger.error(f"报告缺少工作表 '{sheet_name}'，实际: {wb.sheetnames}")
            return {"success": False, "records": [], "rate_limited": False, "fallback_allowed": False, "error": "report missing sheet"}

        ws = wb[sheet_name]
        headers_row = [cell.value for cell in ws[1]]

        # 列名映射（俄语 → 英文）- 只取流量字段
        col_map = {
            "Ваш SKU": "offer_id",
            "Название товара": "offer_name",
            "День": "date",
            "Показы моих товаров, шт.": "shows",
            "Клики по товарам, шт.": "clicks",
            "Добавления в корзину, шт.": "to_cart",
        }

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers_row, row))
            row_dict = {}
            for ru_col, key in col_map.items():
                val = d.get(ru_col)
                if val is None:
                    val = 0
                row_dict[key] = val

            # 过滤无效行
            if not row_dict.get("offer_id") or not row_dict.get("shows"):
                continue

            # 解析日期
            date_str = str(row_dict.get("date") or "")
            if date_str and len(date_str) >= 10:
                parts = date_str[-10:].split("-")
                if len(parts) == 3:
                    row_dict["date"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                else:
                    row_dict["date"] = date_str[-10:]

            records.append(row_dict)

        logger.info(f"  shows-sales 解析完毕: {len(records)} 条记录")
        return {"success": True, "records": records, "rate_limited": False, "fallback_allowed": False, "error": None}

    def _fetch_shows_sales_report(self, campaign_id: int, date_from: str, date_to: str) -> List[dict]:
        """
        内部方法：调用 shows-sales 报告接口（campaign 维度），轮询下载解析，返回流量记录列表。
        注意：此方法为 fallback 用途（仅 businessId 请求失败时使用），420 时直接 raise 不重试。
        """
        import httpx
        import io
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        headers = {"Api-Key": self.client.api_token}

        # Step 1: 发起报告生成
        body = {
            "dateFrom": date_from,
            "dateTo": date_to,
            "campaignId": campaign_id,
            "grouping": "OFFERS"
        }
        resp = httpx.post(
            "https://api.partner.market.yandex.ru/v2/reports/shows-sales/generate",
            headers=headers, json=body, timeout=30
        )
        resp.raise_for_status()
        data = resp.json()
        report_id = data.get("result", {}).get("reportId")
        if not report_id:
            raise Exception(f"shows-sales 生成失败: {data}")

        # Step 2: 轮询等待（最多15分钟）
        for i in range(90):
            time.sleep(10)
            info_resp = httpx.get(
                f"https://api.partner.market.yandex.ru/v2/reports/info/{report_id}",
                headers=headers, timeout=15
            )
            info = info_resp.json()
            status = info.get("result", {}).get("status")
            logger.info(f"  shows-sales poll {i+1}: {status}")
            if status == "DONE":
                break
            if status == "FAILED":
                raise Exception(f"shows-sales 报告生成失败: {info}")
        else:
            raise Exception("shows-sales 报告生成超时（15分钟）")

        # Step 3: 下载并解析 XLSX
        download_url = info["result"]["file"]
        dl_resp = httpx.get(download_url, timeout=60)
        dl_resp.raise_for_status()

        try:
            wb = load_workbook(io.BytesIO(dl_resp.content))
        except InvalidFileException:
            raise Exception("报告文件解析失败（非 XLSX 格式）")

        sheet_name = "Аналитика продаж"
        if sheet_name not in wb.sheetnames:
            raise Exception(f"报告缺少工作表 '{sheet_name}'，实际: {wb.sheetnames}")

        ws = wb[sheet_name]
        headers_row = [cell.value for cell in ws[1]]

        # 列名映射（俄语 → 英文）
        col_map = {
            "Ваш SKU": "offer_id",
            "Название товара": "offer_name",
            "День": "date",
            "Показы моих товаров, шт.": "shows",
            "Клики по товарам, шт.": "clicks",
            "Добавления в корзину, шт.": "to_cart",
        }

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers_row, row))
            row_dict = {}
            for ru_col, key in col_map.items():
                val = d.get(ru_col)
                if val is None:
                    val = 0
                row_dict[key] = val

            # 过滤无效行（无 SKU 或无展示会）
            if not row_dict.get("offer_id") or not row_dict.get("shows"):
                continue

            # 解析日期（格式："25-05-2026" → "2026-05-25"）
            date_str = str(row_dict.get("date") or "")
            if date_str and len(date_str) >= 10:
                parts = date_str[-10:].split("-")
                if len(parts) == 3:
                    row_dict["date"] = f"{parts[2]}-{parts[1]}-{parts[0]}"
                else:
                    row_dict["date"] = date_str[-10:]

            records.append(row_dict)

        logger.info(f"  shows-sales 解析完毕: {len(records)} 条记录")
        return records

    def sync_all(self, history: bool = False) -> dict:
        """
        全量同步

        Args:
            history: 是否同步历史数据（新店铺用）

        Returns:
            各模块同步结果
        """
        results = {}

        # 1. 先同步商品
        logger.info("开始同步商品...")
        results["products"] = self.sync_products(overwrite=True)

        if not results["products"]["success"]:
            return results

        # 2. 同步订单
        logger.info("开始同步订单...")
        results["orders"] = self.sync_orders()

        # 3. 同步库存
        logger.info("开始同步库存...")
        results["inventory"] = self.sync_inventory()

        # 4. 同步广告
        logger.info("开始同步广告...")
        if self.shop.platform == "yandex":
            results["ads"] = {"success": True, "message": "Yandex MVP 暂不支持广告"}
        else:
            results["ads"] = self.sync_ads()

        # 5. 同步产品销售漏斗数据
        logger.info("开始同步产品销售漏斗数据...")
        if self.shop.platform == "yandex":
            results["product_sales"] = self.sync_yandex_product_sales(
                days=self.HISTORY_DAYS if history else 7
            )
        else:
            results["product_sales"] = self.sync_product_sales(
                days=self.HISTORY_DAYS if history else 7
            )

        # 6. 同步关键词统计
        logger.info("开始同步关键词统计...")
        if self.shop.platform == "yandex":
            results["keywords"] = {"success": True, "message": "Yandex MVP 暂不支持关键词"}
        else:
            results["keywords"] = self.sync_keywords(
                days=self.HISTORY_DAYS if history else 7
            )

        # 7. Yandex 同步流量数据（shows-sales 报告）
        if self.shop.platform == "yandex":
            logger.info("开始同步 Yandex 流量数据...")
            results["traffic"] = self.sync_yandex_traffic()

        logger.info("全量同步完成!")
        return results

    def sync_keywords(self, days: Optional[int] = 7) -> dict:
        """同步关键词统计数据"""
        from app.services.sync import SyncService as OldSyncService
        old_sync = OldSyncService(self.db, self.shop)
        return old_sync.sync_keyword_stats(days=days)


def calculate_order_profit(db: Session, order: Order) -> Order:
    """计算订单利润"""
    items = db.query(OrderItem).filter(OrderItem.order_id == order.id).all()

    total_product_cost = 0

    for item in items:
        if item.product_id:
            records = db.query(InventoryRecord).filter(
                InventoryRecord.product_id == item.product_id,
                InventoryRecord.remaining_quantity > 0
            ).order_by(InventoryRecord.inbound_at).all()

            remaining_qty = item.quantity
            item_cost = 0

            for record in records:
                if remaining_qty <= 0:
                    break

                take_qty = min(remaining_qty, record.remaining_quantity)
                item_cost += take_qty * record.product_cost
                record.remaining_quantity -= take_qty
                remaining_qty -= take_qty

            item.product_cost = item_cost
            total_product_cost += item_cost

    order.product_cost = total_product_cost
    order.profit = (
        order.total_amount
        - order.commission
        - order.logistics_fee
        - order.product_cost
        - order.ad_cost
        - order.other_cost
    )

    if order.total_amount > 0:
        order.profit_rate = order.profit / order.total_amount

    return order
